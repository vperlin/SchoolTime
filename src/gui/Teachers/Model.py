from PySide6.QtCore import Qt, QAbstractItemModel, QModelIndex, Slot
from PySide6.QtGui import QColor
import openpyxl
import csv

import data

import logging
LOG = logging.getLogger(__name__)
LOG.setLevel(logging.DEBUG)


def resetting_model(function):
    def function_resetting_model(self, *args, **kwargs):
        self.beginResetModel()
        try:
            return function(self, *args, **kwargs)
        finally:
            self.endResetModel()
    return function_resetting_model


class Model(QAbstractItemModel):

    def __init__(self, parent=None):
        super().__init__(parent)

        self.__subjects = []
        self.__teachers = []

    def subject_by_id(self, iid):
        for sbj in self.__subjects:
            if sbj.iid == iid:
                return sbj
        else:
            raise KeyError(f'Subject: {iid}')

    def subject_by_code(self, code):
        for sbj in self.__subjects:
            if sbj.code == code:
                return sbj
        else:
            raise KeyError(f'Subject: {code}')

    @property
    def idx_teachers(self):
        return self.index(1, 0, QModelIndex())

    @property
    def idx_subjects(self):
        return self.index(0, 0, QModelIndex())

    def index(self, row, column, idx_parent=QModelIndex()):
        if not idx_parent.isValid():
            match row:
                case 0:
                    return self.createIndex(row, column,
                                            data.Kind.SUBJECT.value)
                case 1:
                    return self.createIndex(row, column,
                                            data.Kind.TEACHER.value)
                case _:
                    LOG.warning(f'Invalid row number = {row}')
                    return QModelIndex()
        else:
            match idx_parent.internalId():
                case data.Kind.SUBJECT.value:
                    return self.createIndex(row, column,
                                            data.Kind.SUBJECT.value | 1)
                case data.Kind.TEACHER.value:
                    return self.createIndex(row, column,
                                            data.Kind.TEACHER.value | 1)
                case _:
                    LOG.warning('Unknown index()')
                    return QModelIndex()

    def parent(self, idx):
        if idx.internalId() in {data.Kind.SUBJECT.value,
                                data.Kind.TEACHER.value}:
            return QModelIndex()
        else:
            match idx.internalId() & data.Kind.KIND.value:
                case data.Kind.SUBJECT.value:
                    return self.createIndex(0, 0, data.Kind.SUBJECT.value)
                case data.Kind.TEACHER.value:
                    return self.createIndex(1, 0, data.Kind.TEACHER.value)
                case _:
                    LOG.warning('Unknown parent')
                    return QModelIndex()

    def rowCount(self, idx_parent=QModelIndex()):
        if not idx_parent.isValid():
            return 2
        match idx_parent.internalId():
            case data.Kind.SUBJECT.value:
                return len(self.__subjects)
            case data.Kind.TEACHER.value:
                return len(self.__teachers)
            case _:
                return 0

    def columnCount(self, idx_parent=QModelIndex()):
        if not idx_parent.isValid():
            return 1
        match idx_parent.internalId():
            case data.Kind.SUBJECT.value:
                return 4
            case data.Kind.TEACHER.value:
                return 9
            case _:
                return 0

    def background_data(self, idx):
        match idx.internalId():
            case data.Kind.SUBJECT.value:
                return None
            case data.Kind.TEACHER.value:
                return None
            case _:
                match idx.internalId() & data.Kind.KIND.value:
                    case data.Kind.TEACHER.value:
                        dt = self.__teachers[idx.row()]
                        if dt.iid is None:
                            return QColor('lightgreen')
                        else:
                            return None
                    case data.Kind.SUBJECT.value:
                        dt = self.__subjects[idx.row()]
                        if dt.iid is None:
                            return QColor('lightgreen')
                        else:
                            return None
                    case _:
                        return None

    def display_data(self, idx):
        match idx.internalId():
            case data.Kind.SUBJECT.value:
                return self.tr('Subjects')
            case data.Kind.TEACHER.value:
                return self.tr('Teachers')
            case _:
                match idx.internalId() & data.Kind.KIND.value:
                    case data.Kind.TEACHER.value:
                        dt = self.__teachers[idx.row()]
                        return dt[idx.column()]
                    case data.Kind.SUBJECT.value:
                        dt = self.__subjects[idx.row()]
                        return dt[idx.column()]
                    case _:
                        return None

    def data(self, idx, role=Qt.DisplayRole):
        match role:
            case Qt.DisplayRole:
                return self.display_data(idx)
            case Qt.BackgroundRole:
                return self.background_data(idx)
            case _:
                return None

    @Slot()
    def reload(self):
        self.beginResetModel()
        try:

            with data.connect() as cursor:

                cursor.execute('''
                    select iid, code, title, note
                        from subjects ;
                ''')
                self.__subjects = [data.Subject(**x) for x in cursor]

                cursor.execute('''
                    select iid, last_name, first_name, middle_name,
                           phone, email, note, subjects, lead_group
                        from teachers_info1 ;
                ''')
                tt = []
                for x in cursor:
                    sbj = []
                    for sid in x['subjects']:
                        sbj.append(self.subject_by_id(sid))
                    x['subjects'] = sbj
                    tt.append(data.Teacher(**x))
                self.__teachers = tt

            # @TODO Возможны дополнительные операции

        finally:
            self.endResetModel()

    @Slot()
    def save(self):
        self.beginResetModel()
        try:

            with data.connect() as cursor:
                # Сохраняем предметы
                for subj in self.__subjects:
                    subj.save(cursor)

                # Сохраняем учителей
                for teac in self.__teachers:
                    teac.save(cursor)

            pass
        finally:
            self.endResetModel()

    @resetting_model
    def load_teachers_csv(self, path):
        with path.open('rt', encoding='utf-8') as src:
            rdr = csv.reader(src)
            for line in rdr:
                fio, phone, email, subjects = line
                if not email:
                    email = None
                if not phone:
                    phone = None
                last_name, first_name, middle_name = fio.split()
                if subjects:
                    subj = subjects.split(',')
                    sbj = []
                    for scode in subj:
                        try:
                            s = self.subject_by_code(scode)
                            sbj.append(s)
                        except KeyError:
                            new_subject = data.Subject(code=scode)
                            self.__subjects.insert(0, new_subject)
                            sbj.append(new_subject)
                    subj = sbj
                else:
                    subj = []
                t = data.Teacher(last_name=last_name, first_name=first_name,
                                 middle_name=middle_name, phone=phone,
                                 email=email, subjects=subj)
                self.__teachers.insert(0, t)

    @resetting_model
    def load_teachers_xlsx(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb['teachers']
        for r in range(1, 10000):
            name = ws.cell(row=r, column=1).value
            if name is None: continue
            name = name.strip()
            if not name: continue
            last_name, first_name, *middle_name = name.split()
            if middle_name:
                middle_name = middle_name[0]
            else:
                middle_name = None
            phone = ws.cell(row=r, column=2).value
            if phone is not None:
                phone = phone.strip()
                if not phone: phone = None 
            email = ws.cell(row=r, column=3).value
            if email is not None:
                email = email.strip()
                if not email: email = None 
            subj = ws.cell(row=r, column=4).value
            if subj is not None:
                subj = subj.split(',')
                sbj = []
                for scode in subj:
                    try:
                        s = self.subject_by_code(scode)
                        sbj.append(s)
                    except KeyError:
                        new_subject = data.Subject(code=scode)
                        self.__subjects.insert(0, new_subject)
                        sbj.append(new_subject)
                subj = sbj
            else:
                subj = []
            t = data.Teacher(last_name=last_name, first_name=first_name,
                             middle_name=middle_name, phone=phone,
                             email=email, subjects=subj)
            self.__teachers.insert(0, t)
